from flask import Flask, request, send_file, render_template
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__, static_folder='static')

# Função para corrigir as provas e gerar o arquivo final
def corrigir_provas(gabarito_path, respostas_path):
    # Carregar os dados do gabarito e das respostas dos alunos
    gabarito_df = pd.read_excel(gabarito_path)
    respostas_alunos_df = pd.read_excel(respostas_path)

    # Converter o gabarito para um dicionário (Resposta e Valor)
    gabarito_dict = gabarito_df.set_index('Questão')[['Resposta', 'Valor']].to_dict(orient='index')

    # Inicializar listas para armazenar acertos e notas finais
    acertos_list = []
    notas_finais_list = []

    # Iterar pelas linhas dos alunos para corrigir as respostas
    for idx, row in respostas_alunos_df.iterrows():
        acertos = 0
        nota_final = 0

        # Iterar sobre as respostas de cada aluno
        for questao, resposta in row.items():
            if questao != 'ID':  # Ignorar a coluna do nome
                if questao in gabarito_dict:  # Verificar se a questão existe no gabarito
                    resposta_correta = gabarito_dict[questao]['Resposta']
                    valor_questao = gabarito_dict[questao]['Valor']

                    # Comparar respostas e calcular acertos e nota
                    if resposta == resposta_correta:
                        acertos += 1
                        nota_final += valor_questao

        # Adicionar os resultados às listas
        acertos_list.append(acertos)
        notas_finais_list.append(nota_final)

    # Adicionar colunas ao DataFrame original
    respostas_alunos_df['Acertos'] = acertos_list
    respostas_alunos_df['Nota Final'] = notas_finais_list

    # Salvar o DataFrame em um novo arquivo Excel
    output_path = "notas_alunos.xlsx"
    respostas_alunos_df.to_excel(output_path, index=False, sheet_name='Notas')

    # Aplicar formatação (cores verde/vermelha) no Excel
    workbook = load_workbook(output_path)
    sheet = workbook['Notas']
    
    # Preenchimentos para cores de fundo (verde e vermelho)
    green_fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
    red_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")

    # Iterar sobre as células e aplicar as cores
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # min_row=2 pula o cabeçalho
        for col_idx, cell in enumerate(row, start=1):
            # Ignorar colunas extras
            if sheet.cell(row=1, column=col_idx).value not in ['ID', 'Acertos', 'Nota Final']:
                questao = sheet.cell(row=1, column=col_idx).value
                resposta_aluno = cell.value
                resposta_correta = gabarito_dict[questao]['Resposta']

                if resposta_aluno == resposta_correta:
                    cell.fill = green_fill  # Resposta correta
                else:
                    cell.fill = red_fill  # Resposta errada

    # Salvar as alterações
    workbook.save(output_path)
    return output_path

# Rota principal
@app.route("/")
def index():
    return render_template("index.html")

# Rota para receber os arquivos e processá-los
@app.route("/upload", methods=["POST"])
def upload_files():
    gabarito_file = request.files['gabarito']
    respostas_file = request.files['respostas']

    # Salvar os arquivos temporariamente
    gabarito_path = "gabarito_temp.xlsx"
    respostas_path = "respostas_temp.xlsx"
    gabarito_file.save(gabarito_path)
    respostas_file.save(respostas_path)

    try:
        # Corrigir as provas e gerar o arquivo final
        resultado_path = corrigir_provas(gabarito_path, respostas_path)
        return send_file(resultado_path, as_attachment=True)
    except Exception as e:
        return f"Ocorreu um erro: {str(e)}", 500
    finally:
        # Limpar arquivos temporários
        os.remove(gabarito_path)
        os.remove(respostas_path)

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000)